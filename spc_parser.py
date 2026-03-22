"""
SPC Data Visualization Tool - Excel Parser Module

Parses Apple CPK Data Sheet Excel files (.xlsx) with dynamic column detection.
Handles varying column layouts between different vendor files.
Supports both "Data Input" (summary values) and "Raw data" (full profiles) sheets.
"""

import openpyxl
import pandas as pd
import re
from dataclasses import dataclass, field
from collections import OrderedDict
from typing import Optional, Dict, List, Tuple
import io

# ---------------------------------------------------------------------------
# Monkey-patch openpyxl 3.1.x ExternalReference bug
# Strict-OOXML xlsx files have <externalReference r:id="..."/> which openpyxl
# deserializes as {'{ns}id': 'rId13'} but ExternalReference.__init__ expects
# a positional 'id' arg. The fix: make 'id' optional with a default.
# ---------------------------------------------------------------------------
try:
    from openpyxl.packaging.workbook import ExternalReference as _ER
    import inspect as _inspect
    _params = _inspect.signature(_ER.__init__).parameters
    if "id" in _params and _params["id"].default is _inspect.Parameter.empty:
        _ER.__init__.__defaults__ = ("",)
except Exception:
    pass



# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------

@dataclass
class DimensionMeta:
    """Metadata for a single dimension group (e.g. SPC_AA)."""
    dim_no: str                     # e.g. "SPC_AA"
    description: str                # e.g. "landing to E surface height"
    dim_type: str                   # e.g. "Non-Profile Measurement"
    point_numbers: list             # list of point labels per sub-column
    nominal: list                   # nominal value per sub-column
    tol_max: list                   # tolerance max (+) per sub-column
    tol_min: list                   # tolerance min (-) per sub-column
    usl: list                       # upper spec limit per sub-column
    lsl: list                       # lower spec limit per sub-column
    col_indices: list               # 1-based column indices in the sheet
    col_labels: list                # readable column labels for the dataframe


@dataclass
class ParsedFile:
    """Result of parsing a single Excel file."""
    filename: str
    sheet_name: str
    part_number: Optional[str] = None
    part_description: Optional[str] = None
    revision: Optional[str] = None
    factory: Optional[str] = None   # factory/site code (e.g. "FX", "TY")
    dimensions: OrderedDict = field(default_factory=OrderedDict)  # dim_no -> DimensionMeta
    data: Optional[pd.DataFrame] = None  # measurement rows
    meta_columns: list = field(default_factory=list)  # names of metadata columns present


# ---------------------------------------------------------------------------
# Known metadata header names (normalised to lowercase for matching)
# ---------------------------------------------------------------------------

KNOWN_META_HEADERS = {
    "build", "shipment date", "color", "config",
    "vendor serial number", "fabric thickness",
    "2d barcode", "1d barcode", "rm coil", "raw material",
    "start point",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_str(val) -> str:
    """Convert a cell value to a stripped string, or empty string if None."""
    if val is None:
        return ""
    return str(val).strip()


def _safe_num(val):
    """Return a float if numeric, else None."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _is_interval_point(point_label: str) -> bool:
    """Check if a point label is an interval (e.g. 'C11-C12') vs actual (e.g. 'C11')."""
    return bool(re.search(r'C\d+-C\d+', str(point_label)))


# ---------------------------------------------------------------------------
# Core parser
# ---------------------------------------------------------------------------

def _find_dim_no_cell(rows, max_scan_rows=50, max_scan_cols=30):
    """
    Scan the top-left area of a sheet looking for a cell that says "Dim. No."
    (case-insensitive).  Returns (row_1based, col_1based) or (None, None).
    """
    for ri in range(min(max_scan_rows, len(rows))):
        row = rows[ri]
        for ci in range(min(max_scan_cols, len(row))):
            val = row[ci].value
            if val is not None and re.match(r"dim\.?\s*no\.?", str(val).strip(), re.IGNORECASE):
                return ri + 1, ci + 1  # 1-based
    return None, None


def _scan_label_rows(rows, label_col, start_row, end_row):
    """
    Scan a label column for known metadata row labels.
    Returns a dict: normalised_label -> row_1based.
    """
    mapping = {}
    for ri in range(start_row - 1, min(end_row, len(rows))):
        row = rows[ri]
        if label_col - 1 >= len(row):
            continue
        val = row[label_col - 1].value
        if val is None:
            continue
        s = str(val).strip().lower()
        # Normalise common variants
        label_map = {
            "dim. no.": "dim_no", "dim no.": "dim_no", "dim. no": "dim_no",
            "dimension description": "description",
            "dimension type": "dim_type",
            "point number (if applicable)": "point_number",
            "point number": "point_number", "point no.": "point_number",
            "nominal dim.": "nominal", "nominal": "nominal",
            "tol. max. (+)": "tol_max", "tol. max (+)": "tol_max",
            "tol max (+)": "tol_max", "tol max": "tol_max",
            "tol. min. (-)": "tol_min", "tol. min (-)": "tol_min",
            "tol min (-)": "tol_min", "tol min": "tol_min",
            "usl": "usl", "lsl": "lsl",
            "start point": "start_point",
            "sn": "sn",
            "process": "process",
        }
        if s in label_map:
            mapping[label_map[s]] = ri + 1  # 1-based
    return mapping


def _find_data_start(rows, label_col, data_col_start, after_row, max_search=60):
    """
    Find where measurement data rows begin by looking for:
    1. A header row containing "Start Point" or "SN" in any column
    2. First row after metadata with numeric values in data columns
    Returns (header_row_1based_or_None, data_start_row_1based).
    """
    # Strategy 1: look for "Start Point" or "SN" text in the label area
    _search_cols = max(label_col + 1, 20)  # search up to label_col at minimum
    for ri in range(after_row - 1, min(after_row + max_search, len(rows))):
        row = rows[ri]
        for ci in range(min(_search_cols, len(row))):
            val = row[ci].value
            if val is None:
                continue
            s = str(val).strip().lower()
            if s == "start point":
                return ri + 1, ri + 2  # header row, data starts next row
            if s == "sn":
                return ri + 1, ri + 2

    # Strategy 2: find first row with numeric data in dimension columns
    for ri in range(after_row - 1, min(after_row + max_search, len(rows))):
        row = rows[ri]
        num_count = 0
        for ci in range(data_col_start - 1, min(data_col_start + 10, len(row))):
            val = row[ci].value
            if val is not None and _safe_num(val) is not None:
                num_count += 1
        if num_count >= 2:
            return None, ri + 1  # no header row, data starts here

    return None, None


# ---------------------------------------------------------------------------
# Sheet name patterns for non-data sheets (to skip during auto-detection)
# ---------------------------------------------------------------------------

_NON_DATA_SHEET_PATTERNS = [
    re.compile(r"^BoxPlotCht", re.IGNORECASE),
    re.compile(r"^Histo\s+Pivot$", re.IGNORECASE),
    re.compile(r"^Histo\s+Listbox$", re.IGNORECASE),
    re.compile(r"^Histo\s+Curve$", re.IGNORECASE),
]


def _is_non_data_sheet(name: str) -> bool:
    """Return True if the sheet name matches a known non-data pattern."""
    for pat in _NON_DATA_SHEET_PATTERNS:
        if pat.search(name):
            return True
    return False


def _parse_single_sheet(wb, sheet_name: str, sheet_rows: list,
                        dim_no_row: int, dim_no_col: int,
                        filename: str) -> ParsedFile:
    """
    Parse a single sheet that has already been identified as containing
    CPK data (i.e. has a "Dim. No." cell at the given position).

    This is the core parsing logic extracted from parse_excel so it can
    be reused for multi-sheet files.
    """
    label_col = dim_no_col       # column containing row labels
    data_col_start = label_col + 1  # first column of dimension data

    result = ParsedFile(filename=filename, sheet_name=sheet_name)

    # Convenience cell accessor
    def _cell(r, c):
        """Get cell value; r and c are 1-based."""
        if r - 1 < len(sheet_rows):
            row = sheet_rows[r - 1]
            if c - 1 < len(row):
                return row[c - 1].value
        return None

    # ------------------------------------------------------------------
    # 1. File-level metadata (scan near top for "Part Number", etc.)
    # ------------------------------------------------------------------
    for ri in range(min(5, len(sheet_rows))):
        row = sheet_rows[ri]
        for ci, cell in enumerate(row):
            val = cell.value
            if val is None:
                continue
            s = str(val).strip().lower()
            if "part number" in s and ci + 1 < len(row):
                result.part_number = _safe_str(row[ci + 1].value)
            elif "revision" in s and ci + 1 < len(row):
                result.revision = _safe_str(row[ci + 1].value)
            elif "part description" in s and ci + 1 < len(row):
                result.part_description = _safe_str(row[ci + 1].value)

    # ------------------------------------------------------------------
    # 2. Detect metadata row positions by scanning label column
    # ------------------------------------------------------------------
    label_rows = _scan_label_rows(sheet_rows, label_col, dim_no_row, dim_no_row + 40)

    desc_row = label_rows.get("description")
    type_row = label_rows.get("dim_type")
    point_row = label_rows.get("point_number")
    nominal_row = label_rows.get("nominal")
    tol_max_row = label_rows.get("tol_max")
    tol_min_row = label_rows.get("tol_min")
    usl_row = label_rows.get("usl")
    lsl_row = label_rows.get("lsl")

    # ------------------------------------------------------------------
    # 3. Dimension metadata (scan data columns from data_col_start)
    # ------------------------------------------------------------------
    max_col = len(sheet_rows[dim_no_row - 1]) if dim_no_row - 1 < len(sheet_rows) else 0

    col_dim_no = {}
    col_desc = {}
    col_type = {}
    col_point = {}
    col_nominal = {}
    col_tol_max = {}
    col_tol_min = {}
    col_usl = {}
    col_lsl = {}

    for ci in range(data_col_start, max_col + 1):  # 1-based
        raw_val = _safe_str(_cell(dim_no_row, ci))
        if not raw_val:
            continue
        # Some files embed description in the dim_no cell (e.g. "SPC_G\nCombo Flex flatness")
        if "\n" in raw_val:
            parts = raw_val.split("\n", 1)
            dim_no = parts[0].strip()
            embedded_desc = parts[1].strip()
        else:
            dim_no = raw_val
            embedded_desc = ""
        col_dim_no[ci] = dim_no
        col_desc[ci] = _safe_str(_cell(desc_row, ci)) if desc_row else embedded_desc
        col_type[ci] = _safe_str(_cell(type_row, ci)) if type_row else ""
        col_point[ci] = _safe_str(_cell(point_row, ci)) if point_row else ""
        col_nominal[ci] = _safe_num(_cell(nominal_row, ci)) if nominal_row else None
        col_tol_max[ci] = _safe_num(_cell(tol_max_row, ci)) if tol_max_row else None
        col_tol_min[ci] = _safe_num(_cell(tol_min_row, ci)) if tol_min_row else None
        col_usl[ci] = _safe_num(_cell(usl_row, ci)) if usl_row else None
        col_lsl[ci] = _safe_num(_cell(lsl_row, ci)) if lsl_row else None

    # Group by dim_no preserving order
    dim_groups = OrderedDict()
    for ci, dno in col_dim_no.items():
        dim_groups.setdefault(dno, []).append(ci)

    # ------------------------------------------------------------------
    # 3b. Merge numbered sub-dimensions (compact format)
    #     e.g. SPC_HG, SPC_HG.01, SPC_HG.02 ... -> single "SPC_HG" group
    #     Only applies when individual dims are single-column with no
    #     point numbers (i.e. the compact format pattern).
    # ------------------------------------------------------------------
    merged_groups = OrderedDict()  # parent_name -> list of col indices
    merged_descs = {}              # parent_name -> description
    consumed = set()               # dim_nos already merged

    # First pass: find explicit parents with .NNN children
    for dno in list(dim_groups.keys()):
        if dno in consumed:
            continue
        children = []
        for other in dim_groups:
            if other == dno:
                continue
            if re.match(re.escape(dno) + r'\.\d+$', other):
                children.append(other)
        if children:
            all_cols = list(dim_groups[dno])
            for child in children:
                all_cols.extend(dim_groups[child])
                consumed.add(child)
            merged_groups[dno] = all_cols
            merged_descs[dno] = col_desc.get(dim_groups[dno][0], "")
            consumed.add(dno)

    # Second pass: group orphan .NNN siblings with no parent
    # e.g. SPC_1.001, SPC_1.002, ... (no bare SPC_1 exists)
    orphans = OrderedDict()  # prefix -> list of (dno, cols)
    for dno in list(dim_groups.keys()):
        if dno in consumed:
            continue
        m = re.match(r'^(.+)\.\d+$', dno)
        if m:
            prefix = m.group(1)
            orphans.setdefault(prefix, []).append(dno)
        else:
            # Not a numbered dim, keep standalone
            merged_groups[dno] = list(dim_groups[dno])
            consumed.add(dno)

    for prefix, siblings in orphans.items():
        if len(siblings) >= 2:
            # Merge all siblings under the prefix name
            all_cols = []
            for sib in siblings:
                all_cols.extend(dim_groups[sib])
                consumed.add(sib)
            merged_groups[prefix] = all_cols
            merged_descs[prefix] = col_desc.get(dim_groups[siblings[0]][0], "")
        else:
            # Single orphan, keep as-is
            dno = siblings[0]
            merged_groups[dno] = list(dim_groups[dno])
            consumed.add(dno)

    for dno, cols in merged_groups.items():
        desc = col_desc.get(cols[0], "") if dno not in merged_descs else merged_descs[dno]
        dtype = col_type.get(cols[0], "")

        col_labels = []
        point_numbers = []
        for idx, ci in enumerate(cols):
            pt = col_point.get(ci, "")
            if pt:
                point_numbers.append(pt)
                col_labels.append(f"{dno}_{pt}")
            else:
                # Synthesize point label: P0, P1, P2, ...
                syn_pt = f"P{idx}"
                point_numbers.append(syn_pt)
                col_labels.append(f"{dno}_{syn_pt}")

        result.dimensions[dno] = DimensionMeta(
            dim_no=dno,
            description=desc,
            dim_type=dtype,
            point_numbers=point_numbers,
            nominal=[col_nominal.get(ci) for ci in cols],
            tol_max=[col_tol_max.get(ci) for ci in cols],
            tol_min=[col_tol_min.get(ci) for ci in cols],
            usl=[col_usl.get(ci) for ci in cols],
            lsl=[col_lsl.get(ci) for ci in cols],
            col_indices=cols,
            col_labels=col_labels,
        )

    # ------------------------------------------------------------------
    # 4. Find data start row (auto-detect header + data)
    # ------------------------------------------------------------------
    # Search after the last known metadata row
    search_after = max(
        dim_no_row + 10,
        *(v for v in [usl_row, lsl_row, nominal_row, tol_max_row, tol_min_row] if v),
    )
    header_row_idx, data_start_row = _find_data_start(
        sheet_rows, label_col, data_col_start, search_after
    )

    if data_start_row is None:
        # No data rows found; return empty result
        result.data = pd.DataFrame()
        wb.close()
        return result

    # ------------------------------------------------------------------
    # 5. Build metadata column mapping from the header row
    # ------------------------------------------------------------------
    meta_col_map = OrderedDict()
    if header_row_idx is not None and header_row_idx - 1 < len(sheet_rows):
        hrow = sheet_rows[header_row_idx - 1]

        # Read ALL columns from the header row up to (and including) the
        # first measurement data column.  This captures every metadata
        # column regardless of its name — no hardcoded list needed.
        sn_col = None
        for ci, cell in enumerate(hrow, 1):
            val = _safe_str(cell.value).strip()
            if not val:
                continue
            if ci >= data_col_start:
                break  # past the metadata columns — into measurement data
            # Skip if it looks like a dimension label (e.g. "SPC_A")
            if val.upper().startswith("SPC_") or val.upper().startswith("DIM"):
                break
            meta_col_map[val] = ci
            if val.lower() == "sn":
                sn_col = ci
    else:
        # No header row -- check if there's an SN / serial column
        # (compact format has "SN" at label_col-1, serial numbers at label_col-1)
        sn_row_label = label_rows.get("sn")
        if sn_row_label is not None:
            # The SN column is typically one col left of the label col
            sn_col_idx = label_col - 1 if label_col > 1 else 1
            meta_col_map["SN"] = sn_col_idx
        process_row_label = label_rows.get("process")
        if process_row_label is not None:
            meta_col_map["Process"] = 1  # typically col A

    result.meta_columns = list(meta_col_map.keys())

    # ------------------------------------------------------------------
    # 6. Measurement data (from data_start_row onward)
    # ------------------------------------------------------------------
    # Determine which column to use for the "is row populated?" check
    # Prefer "Start Point" or "SN", fallback to first dimension column
    check_col = None
    if "Start Point" in meta_col_map:
        check_col = meta_col_map["Start Point"]
    elif "SN" in meta_col_map:
        check_col = meta_col_map["SN"]

    records = []
    for ri in range(data_start_row - 1, len(sheet_rows)):
        row = sheet_rows[ri]
        rec = {}

        # Metadata columns
        for name, ci in meta_col_map.items():
            if ci - 1 < len(row):
                rec[name] = row[ci - 1].value
            else:
                rec[name] = None

        # Skip empty rows: check sentinel column or look for numeric data
        if check_col is not None:
            if check_col - 1 < len(row):
                sentinel = row[check_col - 1].value
            else:
                sentinel = None
            if sentinel is None:
                continue
        else:
            # No sentinel: check if row has any numeric data in dim columns
            has_data = False
            for dno, dmeta in result.dimensions.items():
                for ci_d in dmeta.col_indices[:3]:
                    if ci_d - 1 < len(row) and _safe_num(row[ci_d - 1].value) is not None:
                        has_data = True
                        break
                if has_data:
                    break
            if not has_data:
                continue

        # Measurement columns
        for dno, dmeta in result.dimensions.items():
            for ci, label in zip(dmeta.col_indices, dmeta.col_labels):
                if ci - 1 < len(row):
                    rec[label] = row[ci - 1].value
                else:
                    rec[label] = None

        records.append(rec)

    result.data = pd.DataFrame(records)

    # Convert Shipment Date to datetime if present
    if "Shipment Date" in result.data.columns:
        result.data["Shipment Date"] = pd.to_datetime(
            result.data["Shipment Date"], errors="coerce"
        )

    # ------------------------------------------------------------------
    # 7. Detect factory / site code
    # ------------------------------------------------------------------
    if "Vendor Serial Number" in result.data.columns:
        vsn_vals = result.data["Vendor Serial Number"].dropna().astype(str)
        if len(vsn_vals) > 0:
            most_common = vsn_vals.mode()
            if len(most_common) > 0:
                result.factory = str(most_common.iloc[0]).strip()

    # Fallback: extract factory prefix from SN column (e.g. "FJS..." -> "FJS")
    if not result.factory and "SN" in result.data.columns:
        sn_vals = result.data["SN"].dropna().astype(str)
        if len(sn_vals) > 0:
            first_sn = sn_vals.iloc[0]
            m = re.match(r'^([A-Z]{2,4})', first_sn)
            if m:
                result.factory = m.group(1)

    # Fallback: try to extract factory from filename (e.g. "FX_K116_...")
    if not result.factory:
        name_parts = filename.split("_")
        if name_parts and re.match(r'^[A-Z]{2,4}$', name_parts[0]):
            result.factory = name_parts[0]

    return result


def _open_workbook(file_or_path):
    """Open workbook and return (wb, filename).

    Uses keep_links=False to skip external references.
    If openpyxl returns 0 sheets (strict-OOXML bug), converts the file
    to transitional OOXML in-memory using zipfile XML namespace rewrite.
    """
    if isinstance(file_or_path, (str,)):
        filename = file_or_path.rsplit("/", 1)[-1].rsplit("\\", 1)[-1]
        wb = openpyxl.load_workbook(file_or_path, data_only=True, read_only=True, keep_links=False)
        if not wb.sheetnames:
            wb.close()
            wb = _open_strict_ooxml(file_or_path)
    else:
        filename = getattr(file_or_path, "name", "uploaded_file")
        file_or_path.seek(0)
        wb = openpyxl.load_workbook(file_or_path, data_only=True, read_only=True, keep_links=False)
        if not wb.sheetnames:
            wb.close()
            file_or_path.seek(0)
            wb = _open_strict_ooxml(file_or_path)
    return wb, filename


def _open_strict_ooxml(file_or_path):
    """Convert strict-OOXML xlsx to transitional namespace so openpyxl can read it.

    openpyxl 3.1.x cannot parse files using the strict OOXML namespace
    (http://purl.oclc.org/ooxml/...). This rewrites the XML namespaces
    in-memory to the transitional ones that openpyxl understands.
    """
    import zipfile
    _STRICT_TO_TRANSITIONAL = {
        "http://purl.oclc.org/ooxml/spreadsheetml/main":
            "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "http://purl.oclc.org/ooxml/officeDocument/relationships":
            "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "http://purl.oclc.org/ooxml/drawingml/main":
            "http://schemas.openxmlformats.org/drawingml/2006/main",
        "http://purl.oclc.org/ooxml/drawingml/chart":
            "http://schemas.openxmlformats.org/drawingml/2006/chart",
        "http://purl.oclc.org/ooxml/drawingml/spreadsheetDrawing":
            "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
        "http://purl.oclc.org/ooxml/officeDocument/relationships/worksheet":
            "http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet",
        "http://purl.oclc.org/ooxml/officeDocument/relationships/sharedStrings":
            "http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings",
        "http://purl.oclc.org/ooxml/officeDocument/relationships/styles":
            "http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles",
        "http://purl.oclc.org/ooxml/officeDocument/relationships/theme":
            "http://schemas.openxmlformats.org/officeDocument/2006/relationships/theme",
        "http://purl.oclc.org/ooxml/officeDocument/relationships/externalLink":
            "http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLink",
    }

    buf_in = io.BytesIO()
    if isinstance(file_or_path, str):
        with open(file_or_path, "rb") as f:
            buf_in.write(f.read())
    else:
        file_or_path.seek(0)
        buf_in.write(file_or_path.read())
    buf_in.seek(0)

    buf_out = io.BytesIO()
    with zipfile.ZipFile(buf_in, "r") as zin, zipfile.ZipFile(buf_out, "w") as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.endswith((".xml", ".rels")):
                text = data.decode("utf-8", errors="replace")
                # Also strip conformance="strict" attribute
                text = text.replace(' conformance="strict"', "")
                for strict_ns, trans_ns in _STRICT_TO_TRANSITIONAL.items():
                    text = text.replace(strict_ns, trans_ns)
                data = text.encode("utf-8")
            zout.writestr(item, data)

    buf_out.seek(0)
    import warnings
    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        return openpyxl.load_workbook(buf_out, data_only=True, read_only=True, keep_links=False)


def parse_excel(file_or_path, sheet_name: str = "Raw data") -> ParsedFile:
    """
    Parse a vendor CPK Excel file and return structured data.

    Auto-detects sheet layout by scanning for "Dim. No." marker cells.
    Works with any sheet name and column/row arrangement.

    Parameters
    ----------
    file_or_path : str or file-like
        Path to an .xlsx file, or an in-memory file object (e.g. from
        Streamlit's file_uploader).
    sheet_name : str
        Hint for which sheet to read. If the exact name or known aliases
        are not found, all sheets are scanned for CPK data layout.

    Returns
    -------
    ParsedFile
    """
    wb, filename = _open_workbook(file_or_path)

    # ----- Auto-detect which sheet(s) contain CPK data -----
    _SHEET_ALIASES = {
        "Raw data": ["Raw data", "Raw Data", "raw data", "PP data", "PP"],
        "Data Input": ["Data Input", "data input", "Data input"],
    }
    candidate_sheets = []
    # 1. Exact match
    if sheet_name in wb.sheetnames:
        candidate_sheets.append(sheet_name)
    # 2. Aliases
    for alias in _SHEET_ALIASES.get(sheet_name, []):
        if alias in wb.sheetnames and alias not in candidate_sheets:
            candidate_sheets.append(alias)
    # 3. Case-insensitive match
    lower_target = sheet_name.lower()
    for s in wb.sheetnames:
        if s.lower() == lower_target and s not in candidate_sheets:
            candidate_sheets.append(s)
    # 4. ALL remaining sheets (auto-detect mode)
    for s in wb.sheetnames:
        if s not in candidate_sheets:
            candidate_sheets.append(s)

    # Try each candidate; pick first sheet that has "Dim. No." marker
    for sn in candidate_sheets:
        ws = wb[sn]
        rows = list(ws.rows)
        r, c = _find_dim_no_cell(rows)
        if r is not None:
            result = _parse_single_sheet(wb, sn, rows, r, c, filename)
            wb.close()
            return result

    available = ", ".join(wb.sheetnames)
    wb.close()
    raise ValueError(
        f"No CPK data found in any sheet. Available sheets: {available}"
    )


def parse_excel_multi(file_or_path, sheet_name: str = "Raw data") -> List[ParsedFile]:
    """
    Parse a vendor CPK Excel file and return a list of ParsedFile objects.

    If the requested sheet (e.g. "Raw data") exists, returns a single-element
    list (backward compatible).  If it does not exist, auto-detects ALL data
    sheets by scanning for "Dim. No." markers, skipping known non-data sheets
    (BoxPlotCht*, Histo Pivot, Histo Listbox, Histo Curve).

    This is the preferred entry point for the app layer when a single uploaded
    file may contain multiple data sheets.

    Parameters
    ----------
    file_or_path : str or file-like
        Path to an .xlsx file, or an in-memory file object.
    sheet_name : str
        Preferred sheet name hint (default "Raw data").

    Returns
    -------
    list[ParsedFile]
    """
    wb, filename = _open_workbook(file_or_path)

    # ----- Check for preferred sheet first -----
    _SHEET_ALIASES = {
        "Raw data": ["Raw data", "Raw Data", "raw data", "PP data", "PP"],
        "Data Input": ["Data Input", "data input", "Data input"],
    }
    preferred_names = []
    if sheet_name in wb.sheetnames:
        preferred_names.append(sheet_name)
    for alias in _SHEET_ALIASES.get(sheet_name, []):
        if alias in wb.sheetnames and alias not in preferred_names:
            preferred_names.append(alias)
    lower_target = sheet_name.lower()
    for s in wb.sheetnames:
        if s.lower() == lower_target and s not in preferred_names:
            preferred_names.append(s)

    # If a preferred sheet exists and has data, return just that (classic path)
    for sn in preferred_names:
        ws = wb[sn]
        rows = list(ws.rows)
        r, c = _find_dim_no_cell(rows)
        if r is not None:
            result = _parse_single_sheet(wb, sn, rows, r, c, filename)
            wb.close()
            return [result]

    # ----- No preferred sheet found: scan all sheets for data -----
    results = []
    all_sheet_names = list(wb.sheetnames)  # capture before closing
    for sn in all_sheet_names:
        if _is_non_data_sheet(sn):
            continue
        ws = wb[sn]
        rows = list(ws.rows)
        r, c = _find_dim_no_cell(rows)
        if r is not None:
            try:
                parsed = _parse_single_sheet(wb, sn, rows, r, c, filename)
                if parsed.data is not None and len(parsed.data) > 0:
                    results.append(parsed)
            except Exception:
                # Skip sheets that fail to parse
                continue

    wb.close()

    if not results:
        raise ValueError(
            f"No CPK data found in any sheet. Available sheets: {', '.join(all_sheet_names)}"
        )

    return results


# ---------------------------------------------------------------------------
# Dimension grouping by description keywords
# ---------------------------------------------------------------------------

def detect_dimension_groups(dimensions: OrderedDict) -> Dict[str, List[str]]:
    """
    Auto-detect dimension groups by analysing description keywords.

    Returns a dict of group_display_name -> list of dim_no strings.
    Groups dimensions that share a common keyword in their description.
    Dimensions without a matching keyword are placed in individual groups.
    Also includes an "All dimensions" pseudo-group.
    """
    # Build keyword -> list of dim_nos mapping
    keyword_map: Dict[str, List[Tuple[str, str]]] = {}  # keyword -> [(dim_no, description)]
    ungrouped: List[Tuple[str, str]] = []  # dims with no keyword match

    for dno, dmeta in dimensions.items():
        desc = dmeta.description.lower().strip()
        if not desc:
            ungrouped.append((dno, ""))
            continue

        keyword = _extract_group_keyword(desc)
        if keyword:
            keyword_map.setdefault(keyword, []).append((dno, dmeta.description))
        else:
            ungrouped.append((dno, dmeta.description))

    groups: Dict[str, List[str]] = OrderedDict()

    # Keyword-matched groups (2+ dimensions sharing a keyword)
    for keyword, dim_list in keyword_map.items():
        if len(dim_list) >= 2:
            dim_nos = [d[0] for d in dim_list]
            dim_labels = " / ".join(dim_nos)
            display_keyword = keyword.replace("_", " ").title()
            group_label = f"{display_keyword}: {dim_labels}"
            groups[group_label] = dim_nos
        else:
            # Single-member keyword group -> treat as individual
            ungrouped.extend(dim_list)

    # Individual dimension entries
    for dno, desc in ungrouped:
        label = f"{dno} - {desc}" if desc else dno
        groups[label] = [dno]

    # "All dimensions" pseudo-group
    if len(dimensions) > 1:
        all_dim_nos = list(dimensions.keys())
        groups["All dimensions"] = all_dim_nos

    return groups


def _extract_group_keyword(description: str) -> str:
    """
    Extract a grouping keyword from a dimension description.

    Examples:
        "z straightness of left"       -> "z_straightness"
        "z straightness of front"      -> "z_straightness"
        "flatness of Datum A"          -> "flatness"
        "overall length(outer edge)"   -> "overall_length"
        "half length"                  -> "half_length"
        "half width"                   -> "half_width"
        "landing to E surface height"  -> "landing_height"
    """
    desc = description.lower().strip()

    # Try matching known patterns (most specific first)
    patterns = [
        (r"z\s*straightness", "z_straightness"),
        (r"flatness", "flatness"),
        (r"overall\s*length", "overall_length"),
        (r"half\s*length", "half_length"),
        (r"half\s*width", "half_width"),
        (r"landing.*height", "landing_height"),
        (r"height", "height"),
        (r"thickness", "thickness"),
        (r"gap", "gap"),
        (r"offset", "offset"),
        (r"profile", "profile"),
        (r"straightness", "straightness"),
    ]

    for pattern, keyword in patterns:
        if re.search(pattern, desc):
            return keyword

    return ""


def get_filtered_dim_meta(
    dmeta: DimensionMeta,
    exclude_intervals: bool = True,
) -> Tuple[List[str], List[str], List, List, List]:
    """
    Return filtered lists of (col_labels, point_numbers, nominal, usl, lsl)
    optionally excluding interval points (e.g. "C11-C12").

    Parameters
    ----------
    dmeta : DimensionMeta
    exclude_intervals : bool
        If True, exclude interval-type points like "C11-C12".

    Returns
    -------
    (col_labels, point_numbers, nominal, usl, lsl) -- filtered lists
    """
    col_labels = []
    point_numbers = []
    nominal = []
    usl = []
    lsl = []

    for i, pt in enumerate(dmeta.point_numbers):
        if exclude_intervals and _is_interval_point(pt):
            continue
        col_labels.append(dmeta.col_labels[i])
        point_numbers.append(pt)
        nominal.append(dmeta.nominal[i])
        usl.append(dmeta.usl[i])
        lsl.append(dmeta.lsl[i])

    return col_labels, point_numbers, nominal, usl, lsl


# ---------------------------------------------------------------------------
# Convenience helpers for the app layer
# ---------------------------------------------------------------------------

def get_dimension_options(parsed: ParsedFile) -> list:
    """
    Return a list of (display_label, dim_no) tuples for the dimension selector.
    """
    options = []
    for dno, dmeta in parsed.dimensions.items():
        label = f"{dno} - {dmeta.description}" if dmeta.description else dno
        options.append((label, dno))
    return options


def get_groupable_columns(parsed: ParsedFile) -> list:
    """
    Return the list of metadata column names that can be used for
    X-axis grouping or color-by.
    """
    usable = []
    for name in parsed.meta_columns:
        if name == "Start Point":
            continue
        # Only include columns that actually have data
        if name in parsed.data.columns and parsed.data[name].notna().any():
            usable.append(name)
    return usable
