"""
Sheet Manager — Dimension coverage comparison across uploaded files.

Upload multiple CPK Excel files, pick two file/sheet combos, and see which
dimensions are shared, missing, or fuzzy-matched between them.
"""

import io
import os
import sys
from collections import OrderedDict
from difflib import SequenceMatcher

import streamlit as st
import pandas as pd

_project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _project_root not in sys.path:
    sys.path.insert(0, _project_root)

from spc_parser import parse_excel_multi
from ui_theme import (
    inject_theme, FONT_MONO, FONT_BODY, FONT_HEADING,
    TEXT_PRIMARY, TEXT_SECONDARY, TEXT_MUTED,
    ACCENT, DANGER, SUCCESS, WARNING,
    WHITE, BG_SUBTLE, BORDER, BORDER_LIGHT,
)

# ---------------------------------------------------------------------------
# Page config
# ---------------------------------------------------------------------------
st.set_page_config(page_title="Sheet Manager — SPC", layout="wide", initial_sidebar_state="collapsed")
inject_theme()


# ===================================================================
# HELPER FUNCTIONS
# ===================================================================

def scan_sheets(file_bytes: bytes, filename: str) -> list:
    """Lightweight scanner: detect data sheets, count dims/rows, read part#."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=False, data_only=True)
    _NON_DATA_PREFIXES = ("BoxPlotCht", "Histo ")
    _NON_DATA_EXACT = {"Histo Pivot", "Histo Listbox", "Histo Curve"}
    _NON_DATA_KEYWORDS = {"color", "cosmetic", "waive"}

    results = []
    for sn in wb.sheetnames:
        if sn in _NON_DATA_EXACT or any(sn.startswith(p) for p in _NON_DATA_PREFIXES):
            continue
        if any(kw in sn.lower() for kw in _NON_DATA_KEYWORDS):
            continue

        ws = wb[sn]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        if max_row < 5 or max_col < 5:
            continue

        scan_col_limit = min(max_col, 800)

        # Find "Dim. No." row
        dim_row = None
        dim_label_col = None
        for r in range(1, min(16, max_row + 1)):
            for c in range(1, min(31, scan_col_limit + 1)):
                v = ws.cell(r, c).value
                if v and isinstance(v, str) and v.strip().lower().replace(".", "").replace(" ", "") in ("dimno",):
                    dim_row = r
                    dim_label_col = c
                    break
            if dim_row:
                break

        if dim_row is None:
            continue

        # Read unique dims
        dims = []
        col_limit = min(scan_col_limit + 1, dim_label_col + 700)
        for c in range(dim_label_col + 1, col_limit):
            v = ws.cell(dim_row, c).value
            if v is not None and str(v).strip():
                s = str(v).strip()
                if s.lower() not in ("dim. no.", "dim.no.", "dim no"):
                    dims.append(s)
        unique_dims = list(dict.fromkeys(dims))

        # Part number
        part_number = ""
        for r in range(1, min(6, max_row + 1)):
            for c in range(1, min(31, scan_col_limit + 1)):
                v = ws.cell(r, c).value
                if v and isinstance(v, str) and "part number" in v.lower():
                    pn = ws.cell(r, c + 1).value
                    if pn:
                        part_number = str(pn).strip()
                    break
            if part_number:
                break

        header_rows = dim_row + 30
        data_row_estimate = max(0, max_row - header_rows)

        results.append({
            "filename": filename,
            "sheet_name": sn,
            "dim_count": len(unique_dims),
            "dims_list": unique_dims,
            "dims_set": set(unique_dims),
            "data_rows": data_row_estimate,
            "part_number": part_number,
            "selected": True,
        })

    wb.close()
    return results


# ===================================================================
# MAIN PAGE
# ===================================================================

st.markdown(
    f"<h1 style='font-family:{FONT_HEADING};font-size:1.5rem;font-weight:700;"
    f"color:{TEXT_PRIMARY};margin-bottom:0;'>Sheet Manager</h1>"
    f"<p style='font-size:0.8rem;color:{TEXT_MUTED};margin-top:2px;'>"
    f"Upload files → Pick two sheets → Compare dimension coverage</p>",
    unsafe_allow_html=True,
)

# ---------------------------------------------------------------------------
# File Upload
# ---------------------------------------------------------------------------
uploaded_files = st.file_uploader(
    "Upload CPK Excel files (.xlsx)",
    type=["xlsx"],
    accept_multiple_files=True,
    key="sm_uploader",
)

if not uploaded_files:
    st.info("Upload one or more .xlsx files to begin.")
    st.stop()

file_bytes_map = {}
for uf in uploaded_files:
    raw = uf.read()
    uf.seek(0)
    file_bytes_map[uf.name] = raw

# ---------------------------------------------------------------------------
# Scan all sheets
# ---------------------------------------------------------------------------
@st.cache_data(show_spinner="Scanning sheets...")
def _scan_all(file_data: dict) -> list:
    all_infos = []
    for fname, fbytes in file_data.items():
        infos = scan_sheets(fbytes, fname)
        all_infos.extend(infos)
    return all_infos

sheet_infos = _scan_all(file_bytes_map)

if not sheet_infos:
    st.warning("No data sheets found in uploaded files.")
    st.stop()

if len(sheet_infos) < 2:
    st.info("Need at least 2 sheets across uploaded files to compare.")
    st.stop()


# ---------------------------------------------------------------------------
# Dim Coverage helpers
# ---------------------------------------------------------------------------
def _normalize_dim(name: str) -> str:
    s = name.strip().upper().replace(" ", "")
    for prefix in ("SPC_", "SPC-", "DIM_", "DIM-"):
        if s.startswith(prefix):
            s = s[len(prefix):]
    return s


def _fuzzy_match(name_a: str, name_b: str) -> bool:
    if name_a == name_b:
        return False
    na, nb = _normalize_dim(name_a), _normalize_dim(name_b)
    if na == nb:
        return True
    if na.startswith(nb) or nb.startswith(na):
        return True
    if SequenceMatcher(None, na, nb).ratio() >= 0.82:
        return True
    return False


def _short_label(si):
    stem = os.path.splitext(si["filename"])[0]
    tags = []
    for tok in stem.replace("_", " ").replace("-", " ").split():
        t = tok.strip().upper()
        if t in ("CORR", "POR", "FAI", "IQC", "OQC", "GRR", "CPK", "SHIP"):
            tags.append(t)
    tag = " ".join(tags) if tags else stem[:15]
    return f"{tag} / {si['sheet_name']}"


# ---------------------------------------------------------------------------
# Select which two sheets to compare
# ---------------------------------------------------------------------------
_options = [f"{si['filename']} / {si['sheet_name']}" for si in sheet_infos]

sel_col1, sel_col2 = st.columns(2)
with sel_col1:
    pick_a = st.selectbox("File A", options=_options, index=0, key="sm_cov_a")
with sel_col2:
    default_b = min(1, len(_options) - 1)
    pick_b = st.selectbox("File B", options=_options, index=default_b, key="sm_cov_b")

idx_a = _options.index(pick_a)
idx_b = _options.index(pick_b)
si_a = sheet_infos[idx_a]
si_b = sheet_infos[idx_b]

dims_a = si_a["dims_set"]
dims_b = si_b["dims_set"]
dims_all = sorted(dims_a | dims_b)

label_a = _short_label(si_a)
label_b = _short_label(si_b)

# Compute sets
shared = dims_a & dims_b
only_a = dims_a - dims_b
only_b = dims_b - dims_a

fuzzy_pairs = {}
for d in only_a:
    for d2 in dims_b:
        if _fuzzy_match(d, d2):
            fuzzy_pairs[d] = d2
            break
for d in only_b:
    if d not in fuzzy_pairs.values():
        for d2 in dims_a:
            if _fuzzy_match(d, d2):
                fuzzy_pairs[d] = d2
                break

# ---------------------------------------------------------------------------
# Summary metrics
# ---------------------------------------------------------------------------
mc1, mc2, mc3, mc4 = st.columns(4)
mc1.metric("Total Dims", len(dims_all))
mc2.metric("Shared", len(shared))
mc3.metric("Only in A", len(only_a))
mc4.metric("Only in B", len(only_b))

# ---------------------------------------------------------------------------
# Filter controls
# ---------------------------------------------------------------------------
filter_col1, filter_col2 = st.columns(2)
with filter_col1:
    show_filter = st.radio(
        "Show",
        ["All", "Mismatched", "Fuzzy"],
        horizontal=True,
        key="sm_cov_filter",
    )
with filter_col2:
    search_term = st.text_input(
        "Search",
        placeholder="e.g. SPC_A",
        key="sm_cov_search",
    )

# ---------------------------------------------------------------------------
# Build table rows — sorted: Fuzzy → Missing → OK
# ---------------------------------------------------------------------------
table_rows = []
for dim in dims_all:
    in_a = dim in dims_a
    in_b = dim in dims_b
    if in_a and in_b:
        status = "OK"
    elif dim in fuzzy_pairs:
        status = "Fuzzy"
    else:
        status = "Missing"
    table_rows.append({
        "dim": dim,
        "in_a": in_a,
        "in_b": in_b,
        "status": status,
        "fuzzy_hint": fuzzy_pairs.get(dim, ""),
    })

_status_order = {"Fuzzy": 0, "Missing": 1, "OK": 2}
table_rows.sort(key=lambda r: (_status_order.get(r["status"], 9), r["dim"]))

# Apply filters
if show_filter == "Mismatched":
    table_rows = [r for r in table_rows if r["status"] != "OK"]
elif show_filter == "Fuzzy":
    table_rows = [r for r in table_rows if r["status"] == "Fuzzy"]
if search_term:
    q = search_term.strip().upper()
    table_rows = [r for r in table_rows if q in r["dim"].upper()]

# ---------------------------------------------------------------------------
# Render table
# ---------------------------------------------------------------------------
if not table_rows:
    st.caption("No dimensions match the filter.")
else:
    _th = (
        f"padding:6px 10px;font-size:0.72rem;font-weight:600;color:{TEXT_MUTED};"
        f"border-bottom:2px solid {BORDER};font-family:{FONT_BODY};background:{BG_SUBTLE};"
    )
    hdr = (
        f"<th style='{_th}text-align:left;'>Dimension</th>"
        f"<th style='{_th}text-align:center;'>{label_a}</th>"
        f"<th style='{_th}text-align:center;'>{label_b}</th>"
        f"<th style='{_th}text-align:center;'>Status</th>"
    )

    body = ""
    for r in table_rows:
        s = r["status"]
        if s == "OK":
            sc, sb = SUCCESS, f"{SUCCESS}18"
        elif s == "Fuzzy":
            sc, sb = WARNING, f"{WARNING}18"
        else:
            sc, sb = DANGER, f"{DANGER}18"

        rbg = WHITE if s == "OK" else BG_SUBTLE

        # Cell A
        if r["in_a"]:
            ca = f"<span style='color:{SUCCESS};'>&#10003;</span>"
        elif r["fuzzy_hint"] and r["fuzzy_hint"] in dims_a:
            ca = (f"<span style='color:{WARNING};'>&#10007;</span>"
                  f"<br><span style='font-size:0.62rem;color:{WARNING};"
                  f"font-family:{FONT_MONO};'>has {r['fuzzy_hint']}</span>")
        else:
            ca = f"<span style='color:{DANGER};'>&#10007;</span>"

        # Cell B
        if r["in_b"]:
            cb = f"<span style='color:{SUCCESS};'>&#10003;</span>"
        elif r["fuzzy_hint"] and r["fuzzy_hint"] in dims_b:
            cb = (f"<span style='color:{WARNING};'>&#10007;</span>"
                  f"<br><span style='font-size:0.62rem;color:{WARNING};"
                  f"font-family:{FONT_MONO};'>has {r['fuzzy_hint']}</span>")
        else:
            cb = f"<span style='color:{DANGER};'>&#10007;</span>"

        badge = (
            f"<span style='display:inline-block;padding:1px 8px;border-radius:2px;"
            f"font-size:0.65rem;font-weight:600;font-family:{FONT_BODY};"
            f"color:{sc};background:{sb};border:1px solid {sc}22;'>{s}</span>"
        )

        _td = f"border-bottom:1px solid {BORDER_LIGHT};"
        body += (
            f"<tr style='background:{rbg};'>"
            f"<td style='padding:4px 10px;{_td}font-family:{FONT_MONO};"
            f"font-size:0.72rem;color:{TEXT_PRIMARY};'>{r['dim']}</td>"
            f"<td style='padding:4px 10px;{_td}text-align:center;'>{ca}</td>"
            f"<td style='padding:4px 10px;{_td}text-align:center;'>{cb}</td>"
            f"<td style='padding:4px 10px;{_td}text-align:center;'>{badge}</td>"
            f"</tr>\n"
        )

    st.markdown(
        f"<div style='font-size:0.72rem;font-weight:600;text-transform:uppercase;"
        f"letter-spacing:0.06em;color:{TEXT_MUTED};margin:12px 0 6px;'>"
        f"Dimension Coverage — {len(table_rows)} dims</div>",
        unsafe_allow_html=True,
    )
    st.markdown(
        f"<div style='overflow-x:auto;border:1px solid {BORDER};border-radius:2px;"
        f"max-height:500px;overflow-y:auto;'>"
        f"<table style='width:100%;border-collapse:collapse;'>"
        f"<thead style='position:sticky;top:0;z-index:1;'><tr>{hdr}</tr></thead>"
        f"<tbody>{body}</tbody></table></div>",
        unsafe_allow_html=True,
    )

    # Fuzzy match summary
    fuzzy_rows = [r for r in table_rows if r["status"] == "Fuzzy"]
    if fuzzy_rows:
        st.markdown(
            f"<div style='margin-top:12px;padding:8px 12px;background:{BG_SUBTLE};"
            f"border:1px solid {BORDER_LIGHT};border-left:3px solid {WARNING};border-radius:2px;'>"
            f"<span style='font-size:0.75rem;font-weight:600;color:{WARNING};'>"
            f"Fuzzy Matches ({len(fuzzy_rows)})</span>"
            f"<p style='font-size:0.72rem;color:{TEXT_SECONDARY};margin:4px 0 0;'>"
            f"These may be the same dimension with different naming.</p></div>",
            unsafe_allow_html=True,
        )
        for r in fuzzy_rows:
            hint = r["fuzzy_hint"]
            if hint:
                st.markdown(
                    f"<div style='padding:2px 12px 2px 24px;font-size:0.7rem;'>"
                    f"<span style='font-family:{FONT_MONO};color:{ACCENT};'>{r['dim']}</span>"
                    f" <span style='color:{TEXT_MUTED};'>&harr;</span> "
                    f"<span style='font-family:{FONT_MONO};color:{WARNING};'>{hint}</span></div>",
                    unsafe_allow_html=True,
                )
